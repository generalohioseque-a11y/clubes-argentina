import openpyxl
import json
import os

wb = openpyxl.load_workbook('CLUBES.xlsx')
ws = wb.active

clubs = []
logos_dir = os.path.join(os.path.dirname(__file__), 'LOGOS')

# Get all logo files
available_logos = set()
for file in os.listdir(logos_dir):
    if file.endswith('.png') or file.endswith('.jpg'):
        name = file.rsplit('.', 1)[0]
        available_logos.add(int(name))

# Extract club data
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    if len(row) >= 15:
        club_id = row[0]  # ID
        nombre = row[1]   # Nombre completo
        anio_fundacion = row[2]  # Año de fundación
        division = row[3]  # División
        lat = row[4]      # lat
        long = row[5]     # long
        ubicacion = row[6]  # Ubicación completa
        ciudad = row[7]   # Ciudad
        provincia = row[8]  # Provincia
        liga_afiliada = row[9]  # Liga Afiliada
        activo = row[10]  # Activo o inactivo
        liga_juego = row[11]  # Liga en el juego
        jugable = row[12]  # ¿Jugable?
        estadio = row[13]  # Estadio
        camiseta = row[14]  # Camiseta titular
        
        # Check if data is valid
        if club_id and nombre and lat is not None and long is not None:
            try:
                lat = float(lat)
                long = float(long)
                
                # Check if logo exists
                has_logo = int(club_id) in available_logos
                
                clubs.append({
                    'id': int(club_id),
                    'nombre': str(nombre),
                    'anio_fundacion': anio_fundacion if anio_fundacion else 'N/A',
                    'division': str(division) if division else 'N/A',
                    'ubicacion': str(ubicacion) if ubicacion else 'N/A',
                    'ciudad': str(ciudad) if ciudad else 'N/A',
                    'provincia': str(provincia) if provincia else 'N/A',
                    'liga_afiliada': str(liga_afiliada) if liga_afiliada else 'Ninguna',
                    'activo': str(activo) if activo else 'N/A',
                    'liga_juego': str(liga_juego) if liga_juego else 'N/A',
                    'jugable': str(jugable) if jugable else 'N/A',
                    'estadio': str(estadio) if estadio else 'N/A',
                    'camiseta': str(camiseta) if camiseta else 'N/A',
                    'lat': lat,
                    'long': long,
                    'has_logo': has_logo
                })
            except (ValueError, TypeError):
                pass

# Save to JSON
with open('clubs_data.json', 'w', encoding='utf-8') as f:
    json.dump(clubs, f, ensure_ascii=False, indent=2)

# Save to JS file for standalone HTML
with open('clubs_data.js', 'w', encoding='utf-8') as f:
    f.write('const clubsData = ')
    json.dump(clubs, f, ensure_ascii=False)
    f.write(';')

print(f"Total clubs extracted: {len(clubs)}")
print(f"Clubs with logos: {sum(1 for c in clubs if c['has_logo'])}")
